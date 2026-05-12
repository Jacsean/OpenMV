"""
轮廓分析节点 - 提取和分析图像中的轮廓信息

提供完整的轮廓分析功能，包括：
- 轮廓提取和筛选
- 基础统计（周长、面积、边界框、质心）
- 形状检测（圆形、矩形、直线）
- 数据导出（CSV/Excel/JSON）
- 自动二值化检测
- 标注图像绘制
"""

from shared_libs.node_base import BaseNode, ParameterContainerWidget
import cv2
import numpy as np
import os
import csv
import json
from datetime import datetime


class ContoursAnalysisNode(BaseNode):
    """
    轮廓分析节点
    
    从二值化图像中提取轮廓并进行统计分析。
    
    功能特性：
    - 自动检测并二值化输入图像
    - 可配置的轮廓查找模式和近似方法
    - 基于面积的轮廓筛选
    - 计算周长、面积、边界框、质心等统计信息
    - 形状检测（圆形、矩形、直线）及置信度评估
    - 在图像上按形状类型着色绘制轮廓标注
    
    硬件要求：
    - CPU: 1+ 核心
    - 内存: 1GB+
    - GPU: 不需要
    """
    
    __identifier__ = 'feature_extraction'
    NODE_NAME = '轮廓分析'
    
    # 资源等级声明
    resource_level = "light"
    hardware_requirements = {
        'cpu_cores': 1,
        'memory_gb': 1,
        'gpu_required': False,
        'gpu_memory_gb': 0
    }
    
    def __init__(self):
        super(ContoursAnalysisNode, self).__init__()
        
        # 输入端口
        self.add_input('输入图像', color=(100, 255, 100))
        
        # 输出端口
        self.add_output('标注图像', color=(100, 255, 100))
        self.add_output('轮廓数量', color=(100, 100, 255))
        self.add_output('统计数据', color=(255, 255, 100))
        
        # 创建自定义参数容器控件
        self._param_container = ParameterContainerWidget(self.view, 'contour_params', '')
        
        # 轮廓查找参数
        self._param_container.add_combobox('retrieval_mode', '查找模式',
                                           items=['RETR_EXTERNAL', 'RETR_LIST', 'RETR_CCOMP', 'RETR_TREE'])
        
        self._param_container.add_combobox('approximation_method', '近似方法',
                                           items=['CHAIN_APPROX_NONE', 'CHAIN_APPROX_SIMPLE', 'CHAIN_APPROX_TC89_L1', 'CHAIN_APPROX_TC89_KCOS'])
        
        # 筛选参数
        self._param_container.add_spinbox('min_area', '最小面积', value=100, min_value=1, max_value=999999)
        self._param_container.add_spinbox('max_area', '最大面积', value=100000, min_value=1, max_value=999999)
        
        # 形状检测开关
        self._param_container.add_checkbox('enable_circle_detection', '启用圆形检测', state=True)
        self._param_container.add_spinbox('circle_circularity_threshold', '圆度阈值', value=0.85, min_value=0.0, max_value=1.0, double=True)
        
        self._param_container.add_checkbox('enable_rectangle_detection', '启用矩形检测', state=True)
        self._param_container.add_spinbox('rectangle_fill_ratio_threshold', '填充率阈值', value=0.80, min_value=0.0, max_value=1.0, double=True)
        
        self._param_container.add_checkbox('enable_line_detection', '启用直线检测', state=True)
        self._param_container.add_spinbox('line_straightness_threshold', '直线度阈值', value=0.90, min_value=0.0, max_value=1.0, double=True)
        
        # 显示选项
        self._param_container.add_checkbox('draw_contours', '绘制轮廓', state=True)
        self._param_container.add_spinbox('contour_color_r', 'R', value=0, min_value=0, max_value=255)
        self._param_container.add_spinbox('contour_color_g', 'G', value=255, min_value=0, max_value=255)
        self._param_container.add_spinbox('contour_color_b', 'B', value=0, min_value=0, max_value=255)
        self._param_container.add_spinbox('thickness', '线宽', value=2, min_value=1, max_value=5)
        
        # 数据导出参数
        self._param_container.add_combobox('export_mode', '导出模式',
                                           items=['memory_only', 'csv', 'excel', 'json', 'all'])
        
        self._param_container.add_text_input('export_path', '导出路径')
        self._param_container.add_text_input('filename_prefix', '文件名前缀')
        
        # 设置值变化回调
        self._param_container.set_value_changed_callback(self._on_param_changed)
        
        # 添加到节点
        self.add_custom_widget(self._param_container, tab='properties')
    
    def _on_param_changed(self, name, value):
        """参数值变化回调"""
        self.set_property(name, str(value))
    
    def _get_retrieval_mode(self, mode_str):
        """获取轮廓检索模式"""
        modes = {
            'RETR_EXTERNAL': cv2.RETR_EXTERNAL,      # 只检测最外层轮廓
            'RETR_LIST': cv2.RETR_LIST,              # 检测所有轮廓，不建立层次关系
            'RETR_CCOMP': cv2.RETR_CCOMP,            # 检测所有轮廓，建立两层层次关系
            'RETR_TREE': cv2.RETR_TREE               # 检测所有轮廓，建立完整层次关系
        }
        return modes.get(mode_str, cv2.RETR_EXTERNAL)
    
    def _get_approximation_method(self, method_str):
        """获取轮廓近似方法"""
        methods = {
            'CHAIN_APPROX_NONE': cv2.CHAIN_APPROX_NONE,          # 存储所有点
            'CHAIN_APPROX_SIMPLE': cv2.CHAIN_APPROX_SIMPLE,      # 压缩水平、垂直和对角线段
            'CHAIN_APPROX_TC89_L1': cv2.CHAIN_APPROX_TC89_L1,    # Teh-Chin链近似算法L1
            'CHAIN_APPROX_TC89_KCOS': cv2.CHAIN_APPROX_TC89_KCOS # Teh-Chin链近似算法KCOS
        }
        return methods.get(method_str, cv2.CHAIN_APPROX_SIMPLE)
    
    def _get_project_directory(self):
        """
        获取工程目录路径
        
        Returns:
            str: 工程目录路径，若无法获取则返回当前工作目录
        """
        try:
            # 尝试从节点编辑器获取工程路径
            if hasattr(self, 'graph') and self.graph:
                # 假设graph有get_project_path方法
                if hasattr(self.graph, 'get_project_path'):
                    project_path = self.graph.get_project_path()
                    if project_path:
                        return os.path.dirname(project_path)
        except Exception:
            pass
        
        # 降级方案：使用当前工作目录
        return os.getcwd()
    
    def _generate_export_filename(self, ext, count):
        """
        生成导出文件名
        
        Args:
            ext: 文件扩展名 (.csv/.xlsx/.json)
            count: 轮廓数量
            
        Returns:
            str: 文件名
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        prefix = self._param_container.get_values_dict().get('filename_prefix', '')
        return f"{prefix}{timestamp}_{count}{ext}"
    
    def _get_default_export_path(self, ext, count):
        """
        获取默认导出路径
        
        Args:
            ext: 文件扩展名
            count: 轮廓数量
            
        Returns:
            str: 完整的文件路径
        """
        project_dir = self._get_project_directory()
        
        # 创建 exports 子目录
        export_dir = os.path.join(project_dir, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        # 生成文件名
        filename = self._generate_export_filename(ext, count)
        
        return os.path.join(export_dir, filename)
    
    def _resolve_export_path(self, custom_path, ext, count):
        """
        解析自定义导出路径
        
        Args:
            custom_path: 用户指定的路径（可为相对或绝对）
            ext: 文件扩展名
            count: 轮廓数量
            
        Returns:
            str: 完整的文件路径
        """
        if not custom_path or custom_path.strip() == '':
            return self._get_default_export_path(ext, count)
        
        custom_path = custom_path.strip()
        
        # 若是相对路径，转换为绝对路径
        if not os.path.isabs(custom_path):
            project_dir = self._get_project_directory()
            custom_path = os.path.join(project_dir, custom_path)
        
        # 确保目录存在
        dir_path = os.path.dirname(custom_path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        
        # 若用户未指定扩展名，自动添加
        if not custom_path.lower().endswith(ext):
            # 检查是否已有其他扩展名
            if not any(custom_path.lower().endswith(e) for e in ['.csv', '.xlsx', '.xls', '.json']):
                custom_path += ext
        
        return custom_path
    
    def _check_excel_dependency(self):
        """
        检查Excel导出依赖
        
        Returns:
            bool: 是否可用
        """
        try:
            import openpyxl
            return True
        except ImportError:
            self.log_warning("⚠️ Excel导出需要安装openpyxl库")
            self.log_info("💡 安装命令: pip install openpyxl")
            return False
    
    def _export_to_csv(self, stats_data, file_path):
        """
        导出数据到CSV文件
        
        Args:
            stats_data: 统计数据字典
            file_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            contours = stats_data.get('contours', [])
            
            if not contours:
                self.log_warning("没有数据可导出")
                return False
            
            # CSV表头
            headers = [
                'Index', 'Area', 'Perimeter',
                'Centroid_X', 'Centroid_Y',
                'BBox_X', 'BBox_Y', 'BBox_W', 'BBox_H',
                'Shape_Type', 'Shape_Confidence',
                'Circle_Radius',
                'Rectangle_Width', 'Rectangle_Height', 'Rectangle_Angle',
                'Line_Length', 'Line_Angle'
            ]
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # 写入表头
                writer.writerow(headers)
                
                # 写入数据行
                for stat in contours:
                    row = [
                        stat.get('index', ''),
                        stat.get('area', ''),
                        stat.get('perimeter', ''),
                        stat.get('centroid', {}).get('x', ''),
                        stat.get('centroid', {}).get('y', ''),
                        stat.get('bounding_rect', {}).get('x', ''),
                        stat.get('bounding_rect', {}).get('y', ''),
                        stat.get('bounding_rect', {}).get('w', ''),
                        stat.get('bounding_rect', {}).get('h', ''),
                        stat.get('shape_type', ''),
                        stat.get('shape_confidence', ''),
                        # 圆形数据
                        stat.get('circle', {}).get('radius', '') if stat.get('shape_type') == 'circle' else '',
                        # 矩形数据
                        stat.get('rectangle', {}).get('width', '') if stat.get('shape_type') == 'rectangle' else '',
                        stat.get('rectangle', {}).get('height', '') if stat.get('shape_type') == 'rectangle' else '',
                        stat.get('rectangle', {}).get('angle', '') if stat.get('shape_type') == 'rectangle' else '',
                        # 直线数据
                        stat.get('line', {}).get('length', '') if stat.get('shape_type') == 'line' else '',
                        stat.get('line', {}).get('angle', '') if stat.get('shape_type') == 'line' else ''
                    ]
                    writer.writerow(row)
            
            self.log_success(f"✅ CSV文件已导出: {file_path}")
            return True
        
        except Exception as e:
            self.log_error(f"❌ CSV导出失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _export_to_excel(self, stats_data, file_path):
        """
        导出数据到Excel文件
        
        Args:
            stats_data: 统计数据字典
            file_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            contours = stats_data.get('contours', [])
            total_count = stats_data.get('total_count', 0)
            filtered_count = stats_data.get('filtered_count', 0)
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # Sheet 1: 原始数据
            ws_data = wb.active
            ws_data.title = "原始数据"
            
            # 表头
            headers = [
                'Index', 'Area', 'Perimeter',
                'Centroid_X', 'Centroid_Y',
                'BBox_X', 'BBox_Y', 'BBox_W', 'BBox_H',
                'Shape_Type', 'Shape_Confidence'
            ]
            ws_data.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws_data[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            for stat in contours:
                row = [
                    stat.get('index', ''),
                    stat.get('area', ''),
                    stat.get('perimeter', ''),
                    stat.get('centroid', {}).get('x', ''),
                    stat.get('centroid', {}).get('y', ''),
                    stat.get('bounding_rect', {}).get('x', ''),
                    stat.get('bounding_rect', {}).get('y', ''),
                    stat.get('bounding_rect', {}).get('w', ''),
                    stat.get('bounding_rect', {}).get('h', ''),
                    stat.get('shape_type', ''),
                    stat.get('shape_confidence', '')
                ]
                ws_data.append(row)
            
            # 自动调整列宽
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: 统计摘要
            ws_summary = wb.create_sheet(title="统计摘要")
            
            summary_data = [
                ['指标', '数值'],
                ['总轮廓数', total_count],
                ['筛选后数量', filtered_count],
                ['', ''],
                ['面积统计', ''],
                ['平均面积', round(np.mean([c.get('area', 0) for c in contours]), 2) if contours else 0],
                ['最大面积', round(max([c.get('area', 0) for c in contours]), 2) if contours else 0],
                ['最小面积', round(min([c.get('area', 0) for c in contours]), 2) if contours else 0],
                ['', ''],
                ['形状分布', ''],
            ]
            
            # 统计各形状数量
            shape_counts = {}
            for stat in contours:
                shape_type = stat.get('shape_type', 'unknown')
                shape_counts[shape_type] = shape_counts.get(shape_type, 0) + 1
            
            for shape_type, count in shape_counts.items():
                percentage = round(count / filtered_count * 100, 1) if filtered_count > 0 else 0
                summary_data.append([f"{shape_type}数量", f"{count} ({percentage}%)"])
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            # 设置统计摘要样式
            for cell in ws_summary[1]:
                cell.font = Font(bold=True)
            
            # 保存文件
            wb.save(file_path)
            
            self.log_success(f"✅ Excel文件已导出: {file_path}")
            return True
        
        except ImportError:
            self.log_error("❌ 缺少openpyxl库，无法导出Excel")
            self.log_info("💡 请运行: pip install openpyxl")
            return False
        
        except Exception as e:
            self.log_error(f"❌ Excel导出失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _export_to_json(self, stats_data, file_path):
        """
        导出数据到JSON文件
        
        Args:
            stats_data: 统计数据字典
            file_path: 输出文件路径
            
        Returns:
            bool: 是否成功
        """
        try:
            # 构建完整的JSON数据结构
            export_data = {
                'metadata': {
                    'export_time': datetime.now().isoformat(),
                    'total_count': stats_data.get('total_count', 0),
                    'filtered_count': stats_data.get('filtered_count', 0),
                    'algorithm_version': '1.2.0'
                },
                'contours': stats_data.get('contours', [])
            }
            
            # 写入JSON文件
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            self.log_success(f"✅ JSON文件已导出: {file_path}")
            return True
        
        except Exception as e:
            self.log_error(f"❌ JSON导出失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _export_data(self, stats_data):
        """
        根据配置执行数据导出
        
        Args:
            stats_data: 统计数据字典
        """
        params = self._param_container.get_values_dict()
        export_mode = params.get('export_mode', 'memory_only').lower()
        
        if export_mode == 'memory_only':
            return
        
        custom_path = params.get('export_path', '')
        filtered_count = stats_data.get('filtered_count', 0)
        
        # 定义导出任务列表
        export_tasks = []
        
        if export_mode in ['csv', 'all']:
            ext = '.csv'
            file_path = self._resolve_export_path(custom_path, ext, filtered_count)
            export_tasks.append(('csv', file_path))
        
        if export_mode in ['excel', 'all']:
            # 检查依赖
            if self._check_excel_dependency():
                ext = '.xlsx'
                file_path = self._resolve_export_path(custom_path, ext, filtered_count)
                export_tasks.append(('excel', file_path))
            else:
                self.log_warning("⚠️ Excel导出已跳过（缺少openpyxl依赖）")
        
        if export_mode in ['json', 'all']:
            ext = '.json'
            file_path = self._resolve_export_path(custom_path, ext, filtered_count)
            export_tasks.append(('json', file_path))
        
        # 执行导出任务
        self.log_info(f"📤 开始导出数据 (模式: {export_mode}, 任务数: {len(export_tasks)})")
        
        success_count = 0
        for export_type, file_path in export_tasks:
            if export_type == 'csv':
                if self._export_to_csv(stats_data, file_path):
                    success_count += 1
            elif export_type == 'excel':
                if self._export_to_excel(stats_data, file_path):
                    success_count += 1
            elif export_type == 'json':
                if self._export_to_json(stats_data, file_path):
                    success_count += 1
        
        if success_count > 0:
            self.log_success(f"✅ 数据导出完成 ({success_count}/{len(export_tasks)} 个文件)")
        else:
            self.log_warning("⚠️ 所有导出任务均失败")
    
    def _auto_binarize(self, image):
        """
        自动检测并二值化图像
        
        Args:
            image: 输入图像
            
        Returns:
            tuple: (二值化图像, 是否进行了转换)
        """
        # 转换为灰度图
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            self.log_info("检测到彩色图像，已自动转换为灰度图")
        else:
            gray = image.copy()
        
        # 检查是否已是二值图（只有0和255两个值）
        unique_values = np.unique(gray)
        if len(unique_values) <= 2 and set(unique_values).issubset({0, 255}):
            self.log_info("检测到输入已是二值化图像")
            return gray, False
        
        # 自动二值化（Otsu方法）
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        self.log_warning(f"输入非二值化图像，已自动二值化（阈值={_:.1f}）")
        return binary, True
    
    def _detect_circle(self, contour):
        """
        检测圆形
        
        Args:
            contour: 轮廓点集
            
        Returns:
            dict: 圆形检测结果，包含圆心、半径、圆度、置信度
        """
        try:
            # 计算最小外接圆
            (cx, cy), radius = cv2.minEnclosingCircle(contour)
            
            # 计算理论圆面积
            theoretical_area = np.pi * (radius ** 2)
            
            # 计算实际轮廓面积
            actual_area = cv2.contourArea(contour)
            
            # 计算圆度
            circularity = actual_area / theoretical_area if theoretical_area > 0 else 0
            circularity = min(circularity, 1.0)  # 防止浮点误差
            
            return {
                'center': {'x': int(cx), 'y': int(cy)},
                'radius': int(radius),
                'circularity': round(circularity, 4),
                'confidence': round(circularity, 4)
            }
        except Exception as e:
            self.log_warning(f"圆形检测失败: {e}")
            return None
    
    def _detect_rectangle(self, contour):
        """
        检测矩形
        
        Args:
            contour: 轮廓点集
            
        Returns:
            dict: 矩形检测结果，包含中心、宽高、角度、顶点、填充率、置信度
        """
        try:
            # 计算最小外接矩形
            rect = cv2.minAreaRect(contour)
            (cx, cy), (width, height), angle = rect
            
            # 获取四个顶点
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            
            # 计算矩形面积
            rect_area = width * height
            
            # 计算实际轮廓面积
            actual_area = cv2.contourArea(contour)
            
            # 计算填充率
            fill_ratio = actual_area / rect_area if rect_area > 0 else 0
            fill_ratio = min(fill_ratio, 1.0)
            
            return {
                'center': {'x': int(cx), 'y': int(cy)},
                'width': int(width),
                'height': int(height),
                'angle': round(angle, 2),
                'corners': box.tolist(),
                'fill_ratio': round(fill_ratio, 4),
                'confidence': round(fill_ratio, 4)
            }
        except Exception as e:
            self.log_warning(f"矩形检测失败: {e}")
            return None
    
    def _detect_line(self, contour):
        """
        检测直线
        
        Args:
            contour: 轮廓点集
            
        Returns:
            dict: 直线检测结果，包含起点、终点、角度、长度、直线度、置信度
        """
        try:
            # 拟合直线
            rows, cols = 500, 500  # 默认图像尺寸，用于计算端点
            [vx, vy, x, y] = cv2.fitLine(contour, cv2.DIST_L2, 0, 0.01, 0.01)
            
            # 计算直线角度
            angle = np.arctan2(vy[0], vx[0]) * 180 / np.pi
            
            # 计算轮廓的边界框，用于确定端点
            x_min, y_min, w, h = cv2.boundingRect(contour)
            x_max = x_min + w
            y_max = y_min + h
            
            # 计算直线与边界框的交点作为端点
            # 简化处理：使用轮廓的第一个和最后一个点
            start_point = tuple(contour[0][0])
            end_point = tuple(contour[-1][0])
            
            # 计算长度
            length = np.sqrt((end_point[0] - start_point[0])**2 + 
                           (end_point[1] - start_point[1])**2)
            
            # 计算直线度：点到直线的平均距离
            distances = []
            for point in contour:
                px, py = point[0]
                # 点到直线的距离公式
                dist = abs(vy[0] * (px - x[0]) - vx[0] * (py - y[0]))
                distances.append(dist)
            
            avg_distance = np.mean(distances) if distances else 0
            straightness = 1.0 - (avg_distance / length) if length > 0 else 0
            straightness = max(0.0, min(straightness, 1.0))
            
            return {
                'start_point': {'x': int(start_point[0]), 'y': int(start_point[1])},
                'end_point': {'x': int(end_point[0]), 'y': int(end_point[1])},
                'angle': round(angle, 2),
                'length': round(length, 2),
                'straightness': round(straightness, 4),
                'confidence': round(straightness, 4)
            }
        except Exception as e:
            self.log_warning(f"直线检测失败: {e}")
            return None
    
    def _classify_shape(self, contour, params):
        """
        按优先级判断形状
        
        Args:
            contour: 轮廓点集
            params: 形状检测参数字典
            
        Returns:
            tuple: (shape_type, confidence, shape_data)
        """
        area = cv2.contourArea(contour)
        
        # 过小轮廓跳过形状检测
        if area < 50:
            return ('unknown', 0.0, {})
        
        # 点数过少无法拟合
        if len(contour) < 3:
            return ('unknown', 0.0, {})
        
        # 1. 尝试矩形检测（最高优先级）
        if params['enable_rectangle']:
            rect_result = self._detect_rectangle(contour)
            if rect_result and rect_result['confidence'] >= params['rect_threshold']:
                return ('rectangle', rect_result['confidence'], {'rectangle': rect_result})
        
        # 2. 尝试圆形检测
        if params['enable_circle']:
            circle_result = self._detect_circle(contour)
            if circle_result and circle_result['confidence'] >= params['circle_threshold']:
                return ('circle', circle_result['confidence'], {'circle': circle_result})
        
        # 3. 尝试直线检测
        if params['enable_line']:
            line_result = self._detect_line(contour)
            if line_result and line_result['confidence'] >= params['line_threshold']:
                return ('line', line_result['confidence'], {'line': line_result})
        
        # 4. 未知形状
        return ('unknown', 0.0, {})
    
    def process(self, inputs=None):
        """
        处理节点逻辑
        
        Args:
            inputs: 输入图像
            
        Returns:
            dict: 包含标注图像、轮廓数量和统计数据的字典
        """
        try:
            # Step 1: 验证输入
            if not inputs or len(inputs) == 0 or inputs[0] is None:
                self.log_error("未接收到输入图像")
                return {
                    '标注图像': None,
                    '轮廓数量': 0,
                    '统计数据': {}
                }
            
            image = inputs[0][0] if isinstance(inputs[0], list) else inputs[0]
            
            if image is None or not isinstance(image, np.ndarray):
                self.log_error("输入图像格式错误")
                return {
                    '标注图像': None,
                    '轮廓数量': 0,
                    '统计数据': {}
                }
            
            if image.size == 0:
                self.log_error("输入图像尺寸为0")
                return {
                    '标注图像': None,
                    '轮廓数量': 0,
                    '统计数据': {}
                }
            
            # Step 2: 自动二值化
            binary, was_converted = self._auto_binarize(image)
            
            # Step 3: 读取参数
            params = self._param_container.get_values_dict()
            
            retrieval_mode = self._get_retrieval_mode(params.get('retrieval_mode', 'RETR_EXTERNAL'))
            approximation_method = self._get_approximation_method(params.get('approximation_method', 'CHAIN_APPROX_SIMPLE'))
            
            min_area = float(params.get('min_area', 100))
            max_area = float(params.get('max_area', 100000))
            
            draw_contours = params.get('draw_contours', True)
            contour_color = (
                int(params.get('contour_color_b', 0)),
                int(params.get('contour_color_g', 255)),
                int(params.get('contour_color_r', 0))
            )
            thickness = int(params.get('thickness', 2))
            thickness = max(1, min(5, thickness))  # 限制范围1-5
            
            # 形状检测参数
            enable_circle = params.get('enable_circle_detection', True)
            circle_threshold = float(params.get('circle_circularity_threshold', 0.85))
            enable_rectangle = params.get('enable_rectangle_detection', True)
            rect_threshold = float(params.get('rectangle_fill_ratio_threshold', 0.80))
            enable_line = params.get('enable_line_detection', True)
            line_threshold = float(params.get('line_straightness_threshold', 0.90))
            
            shape_params = {
                'enable_circle': enable_circle,
                'circle_threshold': circle_threshold,
                'enable_rectangle': enable_rectangle,
                'rect_threshold': rect_threshold,
                'enable_line': enable_line,
                'line_threshold': line_threshold
            }
            
            # Step 4: 查找轮廓
            contours, hierarchy = cv2.findContours(binary, retrieval_mode, approximation_method)
            
            total_count = len(contours)
            self.log_info(f"检测到 {total_count} 个轮廓")
            
            if total_count == 0:
                self.log_warning("未检测到任何轮廓，请检查二值化效果")
                return {
                    '标注图像': image.copy(),
                    '轮廓数量': 0,
                    '统计数据': {
                        'total_count': 0,
                        'filtered_count': 0,
                        'contours': []
                    }
                }
            
            # Step 5: 筛选轮廓并计算统计信息
            filtered_contours = []
            contour_stats = []
            
            for idx, contour in enumerate(contours):
                # 计算面积
                area = cv2.contourArea(contour)
                
                # 面积筛选
                if area < min_area or area > max_area:
                    continue
                
                # 计算周长
                perimeter = cv2.arcLength(contour, True)
                
                # 计算边界框
                x, y, w, h = cv2.boundingRect(contour)
                
                # 计算质心
                M = cv2.moments(contour)
                if M['m00'] != 0:
                    cx = int(M['m10'] / M['m00'])
                    cy = int(M['m01'] / M['m00'])
                else:
                    cx, cy = 0, 0
                
                # 形状检测
                shape_type, shape_confidence, shape_data = self._classify_shape(contour, shape_params)
                
                # 记录统计信息
                stat = {
                    'index': idx,
                    'area': round(area, 2),
                    'perimeter': round(perimeter, 2),
                    'bounding_rect': {
                        'x': x,
                        'y': y,
                        'w': w,
                        'h': h
                    },
                    'centroid': {
                        'x': cx,
                        'y': cy
                    },
                    'shape_type': shape_type,
                    'shape_confidence': shape_confidence
                }
                
                # 合并形状数据
                stat.update(shape_data)
                
                contour_stats.append(stat)
                filtered_contours.append(contour)
            
            filtered_count = len(filtered_contours)
            self.log_info(f"筛选后剩余 {filtered_count} 个轮廓")
            
            # Step 6: 绘制标注图像
            if draw_contours and filtered_count > 0:
                # 创建副本用于绘制
                if len(image.shape) == 3:
                    annotated_image = image.copy()
                else:
                    annotated_image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
                
                # 按形状类型分别绘制
                for i, (contour, stat) in enumerate(zip(filtered_contours, contour_stats)):
                    shape_type = stat['shape_type']
                    
                    # 根据形状类型选择颜色
                    if shape_type == 'circle':
                        color = (0, 255, 0)  # 绿色
                        marker_color = (255, 0, 0)  # 红色标记
                    elif shape_type == 'rectangle':
                        color = (255, 0, 0)  # 蓝色
                        marker_color = (255, 255, 0)  # 黄色标记
                    elif shape_type == 'line':
                        color = (255, 0, 255)  # 紫色
                        marker_color = (255, 165, 0)  # 橙色标记
                    else:
                        color = (128, 128, 128)  # 灰色（未知）
                        marker_color = (128, 128, 128)
                    
                    # 绘制轮廓
                    cv2.drawContours(annotated_image, [contour], -1, color, thickness)
                    
                    # 绘制质心标记
                    cx = stat['centroid']['x']
                    cy = stat['centroid']['y']
                    cv2.circle(annotated_image, (cx, cy), 3, marker_color, -1)
                    
                    # 绘制形状特定标注
                    if shape_type == 'circle' and 'circle' in stat:
                        # 绘制圆心和半径
                        circle_data = stat['circle']
                        center = (circle_data['center']['x'], circle_data['center']['y'])
                        radius = circle_data['radius']
                        cv2.circle(annotated_image, center, radius, color, 1)
                    
                    elif shape_type == 'rectangle' and 'rectangle' in stat:
                        # 绘制矩形四顶点
                        rect_data = stat['rectangle']
                        corners = np.array(rect_data['corners'], np.int32)
                        cv2.polylines(annotated_image, [corners], True, color, thickness)
                    
                    elif shape_type == 'line' and 'line' in stat:
                        # 绘制直线
                        line_data = stat['line']
                        start_pt = (line_data['start_point']['x'], line_data['start_point']['y'])
                        end_pt = (line_data['end_point']['x'], line_data['end_point']['y'])
                        cv2.line(annotated_image, start_pt, end_pt, color, thickness)
            else:
                annotated_image = image.copy()
                if len(annotated_image.shape) == 2:
                    annotated_image = cv2.cvtColor(annotated_image, cv2.COLOR_GRAY2BGR)
            
            # Step 7: 构建统计数据
            stats_data = {
                'total_count': total_count,
                'filtered_count': filtered_count,
                'contours': contour_stats
            }
            
            # Step 8: 数据导出（若配置）
            export_mode = params.get('export_mode', 'memory_only').lower()
            if export_mode != 'memory_only':
                self._export_data(stats_data)
            
            self.log_success(f"轮廓分析完成 (总数:{total_count}, 筛选后:{filtered_count})")
            
            return {
                '标注图像': annotated_image,
                '轮廓数量': filtered_count,
                '统计数据': stats_data
            }
            
        except Exception as e:
            self.log_error(f"轮廓分析错误: {e}")
            import traceback
            traceback.print_exc()
            return {
                '标注图像': None,
                '轮廓数量': 0,
                '统计数据': {}
            }